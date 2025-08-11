from __future__ import annotations
import os
import calendar
from dataclasses import dataclass
from datetime import datetime
from typing import List
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

@dataclass
class Row:
    sku: str
    product: str
    brand: str
    value: str  # "si"/"no"

class ExcelReport:
    def __init__(self, when: datetime):
        self.when = when
        self.year = when.year
        self.month = when.month
        self.day = when.day
        self.filename = f"reports/{self.year}/inventory_{self.year}-{self.month:02d}.xlsx"
        os.makedirs(os.path.dirname(self.filename), exist_ok=True)

    def write(self, rows: List[Row]) -> str:
        days_in_month = calendar.monthrange(self.year, self.month)[1]
        cols = ["SKU", "Producto", "Marca"] + [str(d) for d in range(1, days_in_month + 1)]
        day_col = str(self.day)
        df_today = pd.DataFrame(
            [{"SKU": r.sku, "Producto": r.product, "Marca": r.brand, day_col: r.value} for r in rows]
        )

        if os.path.exists(self.filename):
            existing = pd.read_excel(self.filename, sheet_name=calendar.month_name[self.month])
            merged = pd.merge(existing, df_today, on=["SKU", "Producto", "Marca"], how="outer")
            for d in [str(i) for i in range(1, days_in_month + 1)]:
                if d not in merged.columns:
                    merged[d] = None
            merged = merged[cols]
        else:
            merged = df_today
            for d in [str(i) for i in range(1, days_in_month + 1)]:
                if d not in merged.columns:
                    merged[d] = None
            merged = merged[cols]

        with pd.ExcelWriter(self.filename, engine="openpyxl") as writer:
            merged.to_excel(writer, sheet_name=calendar.month_name[self.month], index=False)
            ws = writer.book[calendar.month_name[self.month]]
            first_day_col = 4
            last_col = first_day_col + days_in_month - 1
            max_row = ws.max_row
            for col_idx in range(first_day_col, last_col + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{max_row}",
                    FormulaRule(formula=[f"LOWER({col_letter}2)=\"si\""], fill=GREEN),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{max_row}",
                    FormulaRule(formula=[f"LOWER({col_letter}2)=\"no\""], fill=RED),
                )
        return self.filename
